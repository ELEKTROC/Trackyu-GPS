"""
Script d'assignation des 211 balises non assignées.
Croise les données du fichier Excel de migration avec la DB production.
Génère les requêtes SQL UPDATE pour assigner assigned_client_id et assigned_vehicle_id.
"""
import openpyxl, csv, re

# 1. Load Excel IMEI -> client/plaque mapping
wb = openpyxl.load_workbook(
    r'C:\Users\ADMIN\Desktop\TRACKING\Migrations balises\Vehicle_29-01-2026_11-21-13_PM.xlsx',
    read_only=True
)
ws = wb[wb.sheetnames[0]]
imei_map = {}
for row in ws.iter_rows(min_row=5, values_only=True):
    imei = str(row[4]).strip() if row[4] else ''
    if imei:
        imei_map[imei] = {
            'client': str(row[0] or '').strip(),
            'branche': str(row[1] or '').strip(),
            'plaque': str(row[2] or '').strip(),
            'vehicule': str(row[3] or '').strip()
        }

# 2. Load DB clients
db_clients = {}  # name_upper -> id
db_clients_by_id = {}
with open(r'C:\Users\ADMIN\Desktop\TRACKING\temp_clients_db.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        name_upper = row['name'].strip().upper()
        cid = row['id'].strip()
        db_clients[name_upper] = cid
        db_clients_by_id[cid] = row['name'].strip()

print(f"DB clients loaded: {len(db_clients)}")

# 3. Load DB vehicles
db_vehicles_by_plate = {}  # plate_normalized -> (id, client_id, original_plate)
with open(r'C:\Users\ADMIN\Desktop\TRACKING\temp_vehicles_db.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        vid = row['id'].strip()
        vplate = row['plate'].strip()
        vclient = row['client_id'].strip()
        # Normalize: remove spaces, dashes, underscores, uppercase
        norm = re.sub(r'[\s\-_]', '', vplate).upper()
        db_vehicles_by_plate[norm] = (vid, vclient, vplate)

print(f"DB vehicles loaded: {len(db_vehicles_by_plate)}")

# 4. Unassigned IMEIs
unassigned = [
    '865135060086194','352503092670360','862092069367690','352503092681946','862092069430886',
    '865135060066188','864943045470370','864943044266951','354778342544296','354778341221136',
    '351510091899974','864943045852460','351510092367203','864943046430803','864943043645478',
    '864943044589634','354778340090847','351510092391286','351510092739286','351510091952195',
    '354778340774788','864943048071936','354778340786592','354778340784001','354778340786568',
    '351510091909567','354778340780959','354778340785669','354778340780694','354778341221003',
    '864943046973968','351510093054610','354778341217027','354778341219379','864943044701825',
    '351510091905508','354778341221763','354778341347071','864943043421375','351510092735946',
    '351510092895401','354778341405408','351510092894248','351510091759582','352503092686887',
    '354778341390451','354778341389016','354778341384728','354778341389982','864943043422480',
    '354778341396912','354778341389990','354778341393984','354778341389768','354778341396904',
    '354778342203828','354778342194993','351510092727109','864943044700496','864943043423041',
    '354778342196436','354778341345521','354778342546234','354778342545186','864943043421953',
    '864943044700736','864943045941354','354778342545772','354778342544635','864943048067090',
    '354778342547315','354778343057421','864943045940687','865135060086269','354778343066083',
    '864943043645643','354778343059989','864943042794590','864943042794954','864943041797040',
    '354778343258987','354778343262039','864943044701015','864943044267330','354778343261437',
    '354778343268366','864943044701130','354778343265032','354778343260462','354778343259357',
    '354778343267574','354778343261189','864943043645163','864943043645023','864943043645494',
    '864943043644943','864943043644851','864943044267272','864943043424114','864943043382825',
    '864943044210975','354778343312586','864943044210934','864943044266068','864943044267116',
    '864943044267322','864943044265938','864943044267454','864943044266894','864943044266886',
    '864943043393087','864943043392550','864943043422571','864943045350242','864943044700181',
    '864943047854647','864943047787771','864943048072439','862092069385924','864943044619290',
    '864943044619068','864943045350671','864943045852650','864943045852833','864943045852858',
    '864943045853542','864943045852783','864943045852718','864943045853237','864943045940679',
    '864943045853310','864943045940620','864943045941297','864943043645114','864943045941123',
    '864943045940604','864943043645270','864943044266019','864943045940489','864943045940844',
    '864943045940786','864943045941040','864943045941487','864943045940430','864943046425373',
    '864943046436479','864943046430381','351510092724239','351510092897803','351510092736936',
    '864943046430340','864943046427635','351510092139990','351510091762651','864943045469414',
    '864943045469604','351510091916273','351510092387342','864943045470313','864943045793730',
    '864943046972572','864943046972465','864943046973190','864943046975302','864943046971525',
    '864943046973216','864943046970634','864943045940562','864943046426181','864943046436370',
    '354778341405317','864943045470222','864943047788712','864943047788480','864943042794608',
    '864943044265714','864943048071837','864943048238006','864943048239624','864943048240234',
    '864943048238824','864943048236968','864943048240887','351510091956584','864943047790270',
    '865135060929021','864943048071811','865135060910534','864943048236687','864943048240374',
    '864943048073676','864943048072157','864943047690983','864943047725987','864943048066837',
    '15042020194','864943048068320','864943048070383','864943048067546','1502020166',
    '864943048072447','864943048068171','864943048068601','864943043645551','864943043644729',
    '352478993632455','3524789665244444','864943043645288','864943044589063','865135060066667',
    '351510092739237'
]

stock_labels = ['STOCK', 'STOCK SMT', 'TEST TRACKYU', 'IMPAYES ABJ']

def find_client_id(excel_client_name):
    """Find client ID in DB by name matching."""
    upper = excel_client_name.upper()
    # Exact match
    if upper in db_clients:
        return db_clients[upper]
    # Try removing commas and extra spaces
    norm = upper.replace(',', '').replace('  ', ' ').strip()
    for db_name, db_id in db_clients.items():
        db_norm = db_name.replace(',', '').replace('  ', ' ').strip()
        if db_norm == norm:
            return db_id
    # Try subset matching
    for db_name, db_id in db_clients.items():
        if upper in db_name or db_name in upper:
            return db_id
    # Try first+last name match (some DB names have "LAST, FIRST" format)
    parts = upper.split()
    if len(parts) >= 2:
        reversed_name = f"{parts[-1]}, {' '.join(parts[:-1])}"
        if reversed_name in db_clients:
            return db_clients[reversed_name]
        # Also try "LAST FIRST" -> "LAST, FIRST"
        for db_name, db_id in db_clients.items():
            db_parts = db_name.replace(',', '').split()
            if set(parts) == set(db_parts):
                return db_id
    return None

def find_vehicle_id(excel_plaque):
    """Find vehicle ID in DB by plate matching."""
    if not excel_plaque:
        return None
    # Full normalized match
    norm = re.sub(r'[\s\-_]', '', excel_plaque).upper()
    if norm in db_vehicles_by_plate:
        return db_vehicles_by_plate[norm][0]
    # Try just the plate part (before brand name)
    parts = excel_plaque.strip().split()
    if parts:
        short = re.sub(r'[\s\-_]', '', parts[0]).upper()
        if short in db_vehicles_by_plate:
            return db_vehicles_by_plate[short][0]
        # Try combining first 2 parts
        if len(parts) >= 2:
            combined = re.sub(r'[\s\-_]', '', parts[0] + parts[1]).upper()
            if combined in db_vehicles_by_plate:
                return db_vehicles_by_plate[combined][0]
    return None

# 5. Process all
matched_both = []
matched_client_only = []
no_client = []
skipped = []

for imei in unassigned:
    info = imei_map.get(imei)
    if not info:
        no_client.append((imei, 'NOT IN EXCEL', '', ''))
        continue
    
    excel_client = info['client']
    excel_plaque = info['plaque']
    
    if excel_client in stock_labels:
        skipped.append((imei, excel_client, excel_plaque))
        continue
    
    client_id = find_client_id(excel_client)
    vehicle_id = find_vehicle_id(excel_plaque)
    
    if client_id and vehicle_id:
        matched_both.append((imei, client_id, vehicle_id, excel_client, excel_plaque))
    elif client_id:
        matched_client_only.append((imei, client_id, None, excel_client, excel_plaque))
    else:
        no_client.append((imei, excel_client, excel_plaque, 'CLIENT NOT IN DB'))

print(f"\n=== RESULTATS DU MATCHING ===")
print(f"Client + Véhicule trouvés: {len(matched_both)}")
print(f"Client seul (pas de véhicule): {len(matched_client_only)}")
print(f"Client NON trouvé en DB: {len(no_client)}")
print(f"Skipés (stock/test/impayés): {len(skipped)}")
print(f"Total: {len(matched_both) + len(matched_client_only) + len(no_client) + len(skipped)}")

if no_client:
    print(f"\n--- Clients NON trouvés en DB ({len(no_client)}) ---")
    for item in no_client:
        print(f"  IMEI={item[0]} | client='{item[1]}' | plaque='{item[2]}'")

if matched_client_only:
    print(f"\n--- Client OK mais véhicule non trouvé ({len(matched_client_only)}) ---")
    for item in matched_client_only:
        print(f"  IMEI={item[0]} | {item[3]} -> {item[1]} | plaque='{item[4]}'")

# 6. Generate SQL
sql_lines = []
sql_lines.append("-- Assignation des balises non assignées")
sql_lines.append(f"-- Généré le 2026-02-20")
sql_lines.append(f"-- Total: {len(matched_both)} avec client+véhicule, {len(matched_client_only)} avec client seul")
sql_lines.append("BEGIN;")
sql_lines.append("")

for imei, client_id, vehicle_id, excel_client, excel_plaque in matched_both:
    sql_lines.append(f"-- {excel_client} | {excel_plaque}")
    sql_lines.append(f"UPDATE devices SET assigned_client_id = '{client_id}', assigned_vehicle_id = '{vehicle_id}' WHERE imei = '{imei}' AND (assigned_client_id IS NULL OR assigned_client_id = '');")

sql_lines.append("")
sql_lines.append("-- Client trouvé mais pas de véhicule correspondant")
for imei, client_id, _, excel_client, excel_plaque in matched_client_only:
    sql_lines.append(f"-- {excel_client} | plaque non trouvée: {excel_plaque}")
    sql_lines.append(f"UPDATE devices SET assigned_client_id = '{client_id}' WHERE imei = '{imei}' AND (assigned_client_id IS NULL OR assigned_client_id = '');")

sql_lines.append("")
sql_lines.append("COMMIT;")

sql_content = "\n".join(sql_lines)
sql_path = r'C:\Users\ADMIN\Desktop\TRACKING\assign_devices.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write(sql_content)

print(f"\n=== SQL généré dans: {sql_path} ===")
print(f"Requêtes UPDATE: {len(matched_both) + len(matched_client_only)}")
