"""
Script FINAL d'assignation des 211 balises non assignées.
Inclut les mappings manuels pour les noms qui ne matchent pas automatiquement.
Génère le SQL pour mettre à jour assigned_client_id dans la table devices.
"""
import openpyxl, csv, re, unicodedata

# 1. Load Excel IMEI -> client/plaque mapping
wb = openpyxl.load_workbook(
    r'C:\Users\ADMIN\Desktop\TRACKING\Migrations balises\Vehicle_29-01-2026_11-21-13_PM.xlsx',
    read_only=True
)
ws = wb[wb.sheetnames[0]]
imei_map = {}
for row in ws.iter_rows(min_row=5, values_only=True):
    imei = str(row[4]).strip() if row[4] else ''
    if imei:
        imei_map[imei] = {
            'client': str(row[0] or '').strip(),
            'branche': str(row[1] or '').strip(),
            'plaque': str(row[2] or '').strip(),
        }

# 2. Load DB clients
db_clients_by_name = {}
with open(r'C:\Users\ADMIN\Desktop\TRACKING\temp_clients_db.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        name = row['name'].strip()
        cid = row['id'].strip()
        db_clients_by_name[name.upper()] = cid

# 3. Load DB vehicles
db_vehicles = {}
with open(r'C:\Users\ADMIN\Desktop\TRACKING\temp_vehicles_db.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        vid = row['id'].strip()
        vplate = row['plate'].strip()
        norm = re.sub(r'[\s\-_]', '', vplate).upper()
        db_vehicles[norm] = (vid, vplate)

# 4. Manual name mapping Excel->DB (fuzzy analysis results)
MANUAL_CLIENT_MAP = {
    # Exact/close matches found by fuzzy analysis
    'ABOUDOULAY KARAMOKO': 'CLI-ABJ-00492',
    'ADJOUA GAELLE DESIREE KPANGBA': 'CLI-ABJ-00325',
    'AKA ADOU DENIS': 'CLI-ABJ-01193',
    'AKILEY AVIOLA EDWIGE': 'CLI-ABJ-00521',
    'AMANI DJENEBA COULIBALY': 'CLI-ABJ-00644',
    'AMANI NDRI CLARISSE EPSE BOSSE': 'CLI-ABJ-00591',
    'ASSOUAN ROMAN': 'CLI-ABJ-00875',
    'ASSOUMOU INES EMILIE': 'CLI-ABJ-01101',
    'ATSE FIDEL KAMBO': 'CLI-ABJ-00503',
    'CHAALAN AL ASAAD': 'CLI-ABJ-01309',
    'CISSE MOUSSA': 'CLI-ABJ-01194',
    'COULIBALY NANGOHO': 'CLI-ABJ-00010',
    'DABAH SIKA INNOCENT': 'CLI-ABJ-00729',
    'DENIZ BAYRAM': 'CLI-ABJ-00919',
    'DOMAGNI BARTHELEMY': 'CLI-ABJ-00689',
    'FOFANA LLOSSENI': 'CLI-ABJ-01386',   # FOFANA, LOSSENI
    'FOFANA ROCKYA': 'CLI-ABJ-00884',     # FOFANA, ROCKIA
    'GOSSE NEE AKETE FLORENTINE': 'CLI-ABJ-01293',
    'KANGA ARTHUR ROMEO': 'CLI-ABJ-01027',
    'KOFFI ADJOUMANI KOBENAN MICHEL': 'CLI-ABJ-01261',
    'KOUKA LOUIS': 'CLI-ABJ-01205',
    'KOUMAN YAO AXEL': 'CLI-ABJ-00283',
    'KOUAKOU AWA EPSE COULIBALY': 'CLI-ABJ-01496',
    'LEBRY PATRICK': 'CLI-ABJ-00573',
    'LOUIS KONAN AMANI': 'CLI-ABJ-00441',
    'M. KOUADIO DECHI': 'CLI-ABJ-00326',
    'MAMA EPSE TCHBOZO': 'CLI-ABJ-00138', # MAMA EPSE TCHIBOZO ISLAMIA ADEBAYO
    'MARINA TRAORE': 'CLI-ABJ-00593',
    'MOSSOU MELANIE': 'CLI-ABJ-01243',
    'Nathalie Unternährer': 'CLI-ABJ-00709',  # NATHALIE UNTERNAHRER
    'OBENG-KOFI': 'CLI-ABJ-00035',       # OBENG KOFFI
    'SAMPANAN HASSANE': 'CLI-ABJ-00022',
    'SAWADOGO ADAMA': 'CLI-ABJ-01285',
    'SHANA GROUP': 'CLI-ABJ-01283',       # SHANA GOUP (typo in DB)
    'SOGODOGO NOHOU': 'CLI-ABJ-01383',
    'SORO ZIE': 'CLI-ABJ-01470',
    'SOUHOUNE SERGE ABRAHAM': 'CLI-ABJ-00620',
    'TIOMAN AHOSSIBARTHELEMY': 'CLI-ABJ-00727', # AHOSSI BARTHELEMY TIOMAN
    'TONOUKOUIN MAHOUGnON SANDRINE': 'CLI-ABJ-00687',
    'TOURE MASSAMGBE': 'CLI-ABJ-00849',   # TOURE, MASSANGBE
    'TRAORE HAMED DRISSA': 'CLI-ABJ-01441',
    'VEYSEL CAY': 'CLI-ABJ-01036',
    'WOGNIN INNOCENT': 'CLI-ABJ-01407',
    'YAO AKISSI INGRID EPSE KOUAME': 'CLI-ABJ-01463',
    'YAO DANIEL': 'CLI-ABJ-01271',
    'YEO FOUGNIGUE': 'CLI-ABJ-00468',    # YEO, FOUNGNIGUE
    'ZORKHOT HUSSEIN': 'CLI-ABJ-00685',
    'ZOUMANA SIDIBE': 'CLI-ABJ-01144',
    'ZOUZOUO COUZIGOU SAINT-CYR': 'CLI-ABJ-00835',
    'ZUNON ANDRE JOEL': 'CLI-ABJ-01427',
    # Clients NOT in DB (need to be created or skipped)
    'AUTO24': 'CLI-SMT-00122',
    'BARRO MABOUDJARA': 'CLI-ABJ-01498',
    'BEDA MARIE FRANCE DEBORAH KANGAH': 'CLI-ABJ-01495',
    'CAMARA ABDOUKARIM': 'CLI-ABJ-00641',  # ECFA LOGISTIQUE
    'COULIBALY TIEFING': 'CLI-ABJ-01137',  # 2MGC SARL
    'DAVID EBA': 'CLI-ABJ-01491',
    'GAP': 'CLI-ABJ-01492',
    'GOULIHI JEAN RAYMOND': 'CLI-ABJ-01497',
    'GROUPE BOWL': 'CLI-ABJ-01500',
    'JABER NOUHAD': 'CLI-ABJ-01499',
    'MAHIZAN KOUAME KOUMAN STEVEN': 'CLI-SMT-00979',
    'NEMESIS SERVICE': 'CLI-ABJ-01493',
    'KONDJI YEDE STÉPHANE': 'CLI-ABJ-01290',
    'ORIGINAL SKF': 'CLI-ABJ-00036',
    'PNLP': 'CLI-SMT-00085',            # Compte SAAR, branche PNLP
    'RAEID HOBALLAH': 'CLI-ABJ-01494',
    'SERVAIR (BURGER KING)': 'CLI-SMT-029',
}

# 5. Unassigned IMEIs
unassigned = [
    '865135060086194','352503092670360','862092069367690','352503092681946','862092069430886',
    '865135060066188','864943045470370','864943044266951','354778342544296','354778341221136',
    '351510091899974','864943045852460','351510092367203','864943046430803','864943043645478',
    '864943044589634','354778340090847','351510092391286','351510092739286','351510091952195',
    '354778340774788','864943048071936','354778340786592','354778340784001','354778340786568',
    '351510091909567','354778340780959','354778340785669','354778340780694','354778341221003',
    '864943046973968','351510093054610','354778341217027','354778341219379','864943044701825',
    '351510091905508','354778341221763','354778341347071','864943043421375','351510092735946',
    '351510092895401','354778341405408','351510092894248','351510091759582','352503092686887',
    '354778341390451','354778341389016','354778341384728','354778341389982','864943043422480',
    '354778341396912','354778341389990','354778341393984','354778341389768','354778341396904',
    '354778342203828','354778342194993','351510092727109','864943044700496','864943043423041',
    '354778342196436','354778341345521','354778342546234','354778342545186','864943043421953',
    '864943044700736','864943045941354','354778342545772','354778342544635','864943048067090',
    '354778342547315','354778343057421','864943045940687','865135060086269','354778343066083',
    '864943043645643','354778343059989','864943042794590','864943042794954','864943041797040',
    '354778343258987','354778343262039','864943044701015','864943044267330','354778343261437',
    '354778343268366','864943044701130','354778343265032','354778343260462','354778343259357',
    '354778343267574','354778343261189','864943043645163','864943043645023','864943043645494',
    '864943043644943','864943043644851','864943044267272','864943043424114','864943043382825',
    '864943044210975','354778343312586','864943044210934','864943044266068','864943044267116',
    '864943044267322','864943044265938','864943044267454','864943044266894','864943044266886',
    '864943043393087','864943043392550','864943043422571','864943045350242','864943044700181',
    '864943047854647','864943047787771','864943048072439','862092069385924','864943044619290',
    '864943044619068','864943045350671','864943045852650','864943045852833','864943045852858',
    '864943045853542','864943045852783','864943045852718','864943045853237','864943045940679',
    '864943045853310','864943045940620','864943045941297','864943043645114','864943045941123',
    '864943045940604','864943043645270','864943044266019','864943045940489','864943045940844',
    '864943045940786','864943045941040','864943045941487','864943045940430','864943046425373',
    '864943046436479','864943046430381','351510092724239','351510092897803','351510092736936',
    '864943046430340','864943046427635','351510092139990','351510091762651','864943045469414',
    '864943045469604','351510091916273','351510092387342','864943045470313','864943045793730',
    '864943046972572','864943046972465','864943046973190','864943046975302','864943046971525',
    '864943046973216','864943046970634','864943045940562','864943046426181','864943046436370',
    '354778341405317','864943045470222','864943047788712','864943047788480','864943042794608',
    '864943044265714','864943048071837','864943048238006','864943048239624','864943048240234',
    '864943048238824','864943048236968','864943048240887','351510091956584','864943047790270',
    '865135060929021','864943048071811','865135060910534','864943048236687','864943048240374',
    '864943048073676','864943048072157','864943047690983','864943047725987','864943048066837',
    '15042020194','864943048068320','864943048070383','864943048067546','1502020166',
    '864943048072447','864943048068171','864943048068601','864943043645551','864943043644729',
    '352478993632455','3524789665244444','864943043645288','864943044589063','865135060066667',
    '351510092739237'
]

stock_labels = ['STOCK', 'STOCK SMT', 'TEST TRACKYU', 'IMPAYES ABJ']

def find_vehicle(plaque):
    """Find vehicle ID by plate matching."""
    if not plaque:
        return None
    norm = re.sub(r'[\s\-_]', '', plaque).upper()
    if norm in db_vehicles:
        return db_vehicles[norm][0]
    parts = plaque.strip().split()
    if parts:
        short = re.sub(r'[\s\-_]', '', parts[0]).upper()
        if short in db_vehicles:
            return db_vehicles[short][0]
    return None

# Process all
assigned_client_vehicle = []  # (imei, client_id, vehicle_id, excel_client, plaque)
assigned_client_only = []     # (imei, client_id, excel_client, plaque)
no_client_in_db = []          # (imei, excel_client, plaque)
skipped = []                  # (imei, excel_client)

for imei in unassigned:
    info = imei_map.get(imei)
    if not info:
        continue
    
    excel_client = info['client']
    excel_plaque = info['plaque']
    
    if excel_client in stock_labels:
        skipped.append((imei, excel_client, excel_plaque))
        continue
    
    # Find client ID
    client_id = None
    if excel_client in MANUAL_CLIENT_MAP:
        client_id = MANUAL_CLIENT_MAP[excel_client]
    elif excel_client.upper() in db_clients_by_name:
        client_id = db_clients_by_name[excel_client.upper()]
    
    if not client_id:
        no_client_in_db.append((imei, excel_client, excel_plaque))
        continue
    
    # Find vehicle ID
    vehicle_id = find_vehicle(excel_plaque)
    
    if vehicle_id:
        assigned_client_vehicle.append((imei, client_id, vehicle_id, excel_client, excel_plaque))
    else:
        assigned_client_only.append((imei, client_id, excel_client, excel_plaque))

# Stats
total_assignable = len(assigned_client_vehicle) + len(assigned_client_only)
print("=" * 60)
print("RÉSULTATS FINAUX")
print("=" * 60)
print(f"Client + Véhicule trouvés  : {len(assigned_client_vehicle)}")
print(f"Client seul (véhicule absent): {len(assigned_client_only)}")
print(f"Client NON trouvé en DB     : {len(no_client_in_db)}")
print(f"Skipés (stock/test/impayés) : {len(skipped)}")
print(f"TOTAL                       : {total_assignable + len(no_client_in_db) + len(skipped)}")
print(f"\n→ ASSIGNABLES: {total_assignable} / 211")

# Clients manquants en DB
if no_client_in_db:
    missing_clients = {}
    for imei, cl, pl in no_client_in_db:
        if cl not in missing_clients:
            missing_clients[cl] = 0
        missing_clients[cl] += 1
    print(f"\n--- Clients absents de la table tiers ({len(missing_clients)} clients, {len(no_client_in_db)} balises) ---")
    for cl, count in sorted(missing_clients.items(), key=lambda x: -x[1]):
        print(f"  {count:2d}x  {cl}")

# Generate SQL
sql_lines = []
sql_lines.append("-- ============================================")
sql_lines.append("-- Assignation des balises non assignées")
sql_lines.append("-- Généré le 2026-02-20")
sql_lines.append(f"-- Client+Véhicule: {len(assigned_client_vehicle)}")
sql_lines.append(f"-- Client seul: {len(assigned_client_only)}")
sql_lines.append(f"-- Non-assignables (client absent DB): {len(no_client_in_db)}")
sql_lines.append("-- ============================================")
sql_lines.append("BEGIN;")
sql_lines.append("")

if assigned_client_vehicle:
    sql_lines.append("-- ===== SECTION 1: Client + Véhicule =====")
    for imei, client_id, vehicle_id, excel_client, plaque in assigned_client_vehicle:
        sql_lines.append(f"-- {excel_client} | {plaque}")
        sql_lines.append(f"UPDATE devices SET assigned_client_id = '{client_id}', assigned_vehicle_id = '{vehicle_id}' WHERE imei = '{imei}' AND (assigned_client_id IS NULL OR assigned_client_id = '');")

sql_lines.append("")
sql_lines.append("-- ===== SECTION 2: Client seul (véhicule non trouvé en DB) =====")
for imei, client_id, excel_client, plaque in assigned_client_only:
    sql_lines.append(f"-- {excel_client} | plaque: {plaque}")
    sql_lines.append(f"UPDATE devices SET assigned_client_id = '{client_id}' WHERE imei = '{imei}' AND (assigned_client_id IS NULL OR assigned_client_id = '');")

sql_lines.append("")
sql_lines.append("COMMIT;")

sql_lines.append("")
sql_lines.append("-- ===== NON-ASSIGNABLES (clients absents de la table tiers) =====")
sql_lines.append("-- Ces balises nécessitent la création préalable du client")
for imei, cl, pl in no_client_in_db:
    sql_lines.append(f"-- SKIP: {imei} | {cl} | {pl}")

sql_path = r'C:\Users\ADMIN\Desktop\TRACKING\assign_devices.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write("\n".join(sql_lines))

print(f"\n→ SQL généré: {sql_path}")
print(f"→ {total_assignable} UPDATE prêts à exécuter")
