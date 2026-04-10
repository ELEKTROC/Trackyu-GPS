import openpyxl
import hashlib

wb = openpyxl.load_workbook('SIM MTN.xlsx', data_only=True)
ws = wb.active

# Collect all phone numbers (9-digit integers, need leading 0)
phones = set()
for row in range(2, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=row, column=col).value
        if val is not None:
            num = str(int(val))
            if len(num) == 9:
                phones.add('0' + num)

phones = sorted(phones)
print(f'Total unique MTN SIM numbers: {len(phones)}')

# Generate SQL migration file
lines = []
lines.append('-- Migration: Insert MTN SIM cards from "SIM MTN.xlsx"')
lines.append(f'-- Total: {len(phones)} unique SIM cards')
lines.append(f'-- Operator: MTN')
lines.append(f'-- Generated: 2026-02-21')
lines.append('')
lines.append('BEGIN;')
lines.append('')

for phone in phones:
    sim_id = f'SIM-{phone}'
    phone_intl = f'+225{phone}'
    iccid = f'SIM-{phone}'
    operator = 'MTN'
    status = 'IN_STOCK'
    
    # Escape single quotes (none expected but just in case)
    lines.append(
        f"INSERT INTO sim_cards (id, tenant_id, iccid, phone_number, operator, status, created_at, updated_at) "
        f"VALUES ('{sim_id}', NULL, '{iccid}', '{phone_intl}', '{operator}', '{status}', NOW(), NOW()) "
        f"ON CONFLICT (id) DO NOTHING;"
    )

lines.append('')
lines.append('-- Verify')
lines.append("SELECT count(*) as total_mtn_sims FROM sim_cards WHERE operator = 'MTN';")
lines.append('')
lines.append('COMMIT;')

sql_path = 'backend/migrations/migrate_sim_mtn.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f'SQL migration written to: {sql_path}')
print(f'Rows: {len(phones)} INSERT statements')
