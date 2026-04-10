import openpyxl

wb = openpyxl.load_workbook('SIM MTN.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
ws = wb.active
print(f'Sheet: {ws.title}')
print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
print()
print('=== Headers (row 1) ===')
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    print(f'  Col {col}: {val}')
print()
print('=== First 10 data rows ===')
for row in range(2, min(12, ws.max_row + 1)):
    vals = []
    for col in range(1, ws.max_column + 1):
        vals.append(str(ws.cell(row=row, column=col).value))
    print(f'  Row {row}: ' + ' | '.join(vals))
print()
print(f'=== Last 3 rows ===')
for row in range(max(2, ws.max_row - 2), ws.max_row + 1):
    vals = []
    for col in range(1, ws.max_column + 1):
        vals.append(str(ws.cell(row=row, column=col).value))
    print(f'  Row {row}: ' + ' | '.join(vals))
print()

# Check unique phone numbers
phones = set()
for row in range(2, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=row, column=col).value
        if val is not None:
            s = str(val).strip()
            # Check if it looks like a phone number (digits, 10 chars, starts with 05)
            digits = ''.join(c for c in s if c.isdigit())
            if len(digits) == 10 and digits.startswith('05'):
                phones.add(digits)
            elif len(digits) == 10:
                phones.add(digits)

print(f'Unique phone-like numbers found: {len(phones)}')
if phones:
    sample = sorted(phones)[:10]
    print(f'Sample: {sample}')
