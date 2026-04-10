import openpyxl

wb = openpyxl.load_workbook('SIM MTN.xlsx', data_only=True)
ws = wb.active

# Collect all phone numbers (9-digit integers, need leading 0)
phones = set()
for row in range(2, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=row, column=col).value
        if val is not None:
            # Convert float to int, then to string
            num = str(int(val))
            if len(num) == 9:
                # Add leading 0 to make 10-digit phone
                phones.add('0' + num)

phones = sorted(phones)
print(f'Total unique MTN SIM numbers: {len(phones)}')
print()

# Analyze prefixes
prefixes = {}
for p in phones:
    prefix = p[:3]
    prefixes[prefix] = prefixes.get(prefix, 0) + 1
print('Prefixes distribution:')
for k, v in sorted(prefixes.items()):
    print(f'  {k}: {v}')
print()

print('Sample (first 15):')
for p in phones[:15]:
    print(f'  {p}')
print()
print('Sample (last 5):')
for p in phones[-5:]:
    print(f'  {p}')

# Check for duplicates with existing Orange SIMs in DB
# We'll generate the SQL to check
print(f'\n--- Summary ---')
print(f'Total unique: {len(phones)}')
print(f'Format: +225XXXXXXXXXX (Ivory Coast MTN)')
